# coding=utf-8
import math
import os
import sys
import time
import requests
from openpyxl import load_workbook
import shutil
import datetime
import numpy as np

import ecal.core.core as ecal_core
from ecal.core.subscriber import ProtoSubscriber

import proto_messages.pose_pb2 as pose_pb2

IS_OPEN_ECAL_CALLBAK = False
case_title = None
car_for_load_position_x = None
CAR_MODEL = None

# 设计三个类
'''
一个类用作角度转换处理的操作
一个类用作excel读取写入的操作
一个类用作robotune接口定义的操作
一个ecal类用来调用当前车体方法
'''


# 替换成杨磊的转换方法
# 定义一个坐标系转换的方法
class CoordinateTransformer:
    # 定义特殊角度列表
    SPECIAL_ANGLES = [3.14, -3.14, 1.57, -1.57]
    def __init__(self,x,y,yaw,length):
        self.x = x
        self.y = y
        self.yaw = yaw
        self.length = length
    def yaw_rotation_is_zero(self):
        # 检查yaw是否接近特殊角度
        for special_angle in self.SPECIAL_ANGLES:
            if math.isclose(self.yaw, 3.14, rel_tol=1e-2) or math.isclose(self.yaw, -3.14, rel_tol=1e-2):
                x = self.x - self.length
                return {"x":x, "y":self.y,"yaw":self.yaw}
            elif math.isclose(self.yaw, 1.57, rel_tol=1e-2) or math.isclose(self.yaw, -1.57, rel_tol=1e-2):
                return self.yaw_rotation_for_zero()
            else:
                # 如果不接近特殊角度，调用 rotation_for_z 方法
                return self.rotation_for_z()
    def yaw_rotation_for_zero(self):
        # 如果是90度的话，y与x反写
        initial_point = np.array([self.y,self.x])
        x = initial_point[0] - self.length
        y = initial_point[1]
        translated_point = np.array([y,x])
        return {"x":translated_point[0], "y":translated_point[1],"yaw":self.yaw}

    def rotation_for_z(self):
        initial_point = np.array([self.x,self.y])
        yaw_angle_rad = np.clip(self.yaw, -np.pi, np.pi)

        # 定义旋转中心
        rotation_center = np.array([self.x,self.y])

        
        # 定义二维旋转矩阵（根据yaw角）
        rotation_matrix = np.array([
            [np.cos(yaw_angle_rad), -np.sin(yaw_angle_rad)],
            [np.sin(yaw_angle_rad), np.cos(yaw_angle_rad)]
        ])

        translation_vector_local = np.array([self.length, 0])
        translation_vector_global = np.dot(rotation_matrix, translation_vector_local)

        # 旋转初始点
        rotated_point = rotation_center + np.dot(rotation_matrix, initial_point - rotation_center)

        # 平移旋转后的点
        translated_point = rotated_point + translation_vector_global
        return {"x":translated_point[0],"y":translated_point[1],"yaw":self.yaw}

# 通过ecal发布接收车体信息
def on_message(topic_name, msg, msg_time):
    """
    
    """
    global IS_OPEN_ECAL_CALLBAK
    global car_for_load_position_x
    if not IS_OPEN_ECAL_CALLBAK:
        return
    ecal_get_webot_position = {"x":msg.position.x,"y":msg.position.y,"yaw":msg.orientation.z*msg.orientation.w}
    x,y,yaw = ecal_get_webot_position["x"],ecal_get_webot_position["y"],ecal_get_webot_position["yaw"]
    json_data = CoordinateTransformer(x,y,yaw,car_for_load_position_x).yaw_rotation_is_zero()
    
    if case_title:
        writing_position_in_excel(json_data,file_path)
    else:
        print("获取数据失败")
    IS_OPEN_ECAL_CALLBAK = False

# 根据日期复制一份excel表
def get_copy_new_file(old_file_path,new_file_path,file_name):
    """
    old_file_path: 旧的文件路径
    new_file_path: 新的文件路径
    file_name: 文件名
    return: 新的文件路径

    """
    # 校准传入的文件是否xlsx格式
    current_time = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    old_file_path = f"{old_file_path}{file_name}.xlsx"
    new_file_path = f"{new_file_path}{CAR_MODEL}_{file_name}_{current_time}.xlsx"
    # 校准文件夹是否存在
    if not os.path.exists("/home/visionnav/excel_report"):
        # 如果文件夹不存在，创建文件夹
        os.makedirs("/home/visionnav/excel_report", exist_ok=True)
    shutil.copy(old_file_path, new_file_path)
    print(f"初始化测试用例，文件路径为：{new_file_path}")
    return new_file_path

# 对excel读取卡板参数的操作，并且把当前的托盘参数进行转换后重新复用一份新的卡板参数
def read_excel_lines(file_path):
    """
    file_path: excel文件路径
    """
    try:
        print("开始读取excel文件")
        # 加载工作簿
        workbook = load_workbook(file_path)
        # 获取活动工作表
        sheet = workbook["Sheet1"]

        # 获取首行作为键
        headers = []
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for cell_value in first_row:
            headers.append(cell_value)
        print("坐标反写中")
        # 从第二行开始遍历每一行
        for row_index,row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 插入数据处理方法
            row_dict = dict(zip(headers, row))
            if row_dict["用例标题"] is not None:
                pallet_x = row_dict["托盘x"]
                pallet_y = row_dict["托盘y"]
                pallet_yaw = row_dict["托盘yaw"]
                if pallet_yaw < 0:
                    pallet_length = row_dict["卡板长度"] / 2
                else:
                    pallet_length = -row_dict["卡板长度"] / 2
                new_axis_rotation = CoordinateTransformer(pallet_x,pallet_y,pallet_yaw,pallet_length).yaw_rotation_is_zero()
                sheet[f"E{row_index}"] = new_axis_rotation["y"]
                sheet[f"F{row_index}"] = new_axis_rotation["x"]
        # 对工作簿进行保存
        workbook.save(file_path)
        # 关闭工作簿
        workbook.close()
        print("完成excel卡板入叉面推算")
    except FileNotFoundError:
        print("错误: 文件未找到!")
    except Exception as e:
        print(f"错误: 发生了一个未知错误: {e}")
    return True
def writing_position_in_excel(json_data,file_path):
    """
    json_data: 传入车体的数据
    file_path: excel文件路径
    """
    print("开始写入excel")
    car_position_x = json_data["y"]
    car_position_y = json_data["x"]
    car_position_yaw = json_data["yaw"]
    time_now = datetime.datetime.now()

    wb = load_workbook(file_path)
    ws = wb["Sheet1"]
    print("开始进行车体数据写入")
    for i in ws.iter_rows():
        for cell in i:
            if cell.value == case_title:
                write_x = f"H{cell.row}"
                write_y = f"I{cell.row}"
                write_yaw = f"J{cell.row}"
                write_time = f"P{cell.row}"
                pallet_x = ws[f"E{cell.row}"].value
                pallet_y = ws[f"F{cell.row}"].value
                pallet_yaw = ws[f"G{cell.row}"].value
                
                # if pallet_x - car_position_x > 0.015 or pallet_y - car_position_y > 0.015 or pallet_yaw - car_position_yaw > 0.2:
                #     # 误差过大，记录日志
                #     print("误差过大，记录日志")
                #     if not os.path.exists("/home/visionnav/error_log/"):
                #         os.makedirs("/home/visionnav/error_log/", exist_ok=True)
                #     shutil.copytree('/home/visionnav/log/', f'/home/visionnav/error_log/log_{time_now}')
                ws[write_x] = car_position_x
                ws[write_y] = car_position_y
                ws[write_yaw] = car_position_yaw
                ws[write_time] = time_now
    wb.save(file_path)
    print("保存成功")
    wb.close()

#TODO: 后期整理工作，规整所有robotune接口,执行程序时只调用一个接口，减少重复
    # robotune_url = 'http://127.0.0.1:24311'
    # robotune_tail_get_task_info = robotune_url+'/api/services/task/Agv/GetCurrentTaskInfo'
    # robotune_get_vehicle_position = robotune_url+'/api/services/task/Agv/GetVehiclePosition'
    # robotune_get_debug_status = robotune_url+'/api/services/task/DynamicFlow/GetDebugStatus?taskId='
    # robotune_get_agv_runtime_status = robotune_url+'/api/services/task/VoiceFlow/GetAGVRuntimeStatus'
    # robotune_get_robot_nodes = robotune_url+'/api/services/pm/CommonParameter/GetRootNodes'
    # robotune_get_all_show_nodes = robotune_url+'/api/services/pm/CommonParameter/GetAllShowNodes?uuid=uuid&key=key'


if __name__ == "__main__":
    # # 执行脚本前会默认清除日志
    # if os.path.exists("/home/visionnav/log"):
    #     print("清空AGV日志")
    #     shutil.rmtree("/home/visionnav/log")
    # 设置robotune接口地址
    robotune_url = 'http://127.0.0.1:24311'
    # id = "53"
    # robotune_task_GetFlowInfo = robotune_url+'/api/services/task/DynamicFlow/GetFlowInfo?id='+id
    # 读取robotune接口定义
    robotune_tail_get_task_info = robotune_url+'/api/services/task/Agv/GetCurrentTaskInfo'
    robotune_get_vehicle_position = robotune_url+'/api/services/task/Agv/GetVehiclePosition'
    robotune_get_debug_status = robotune_url+'/api/services/task/DynamicFlow/GetDebugStatus?taskId='
    robotune_get_agv_runtime_status = robotune_url+'/api/services/task/VoiceFlow/GetAGVRuntimeStatus'

    # 读取当前车体配置的通用参数,查找到子节点ID放入到UUID中，再根据关键字自行数据匹配
    robotune_get_robot_nodes = robotune_url+'/api/services/pm/CommonParameter/GetRootNodes'

    response_getrobotnodes = requests.get(robotune_get_robot_nodes).json()

    # 读取静态参数
    robot_nodes_list = [i["uuid"] for i in response_getrobotnodes["result"]["childNodes"] if i["nodeName"] == "静态参数"]
    # 读取关键字为LoadPositionX的节点
    robotune_get_all_show_nodes = robotune_url+'/api/services/pm/CommonParameter/GetAllShowNodes?'+'UUID='+robot_nodes_list[0]+'&Key=LoadPositionX'
    robotune_get_all_show_nodes_car_model = robotune_url+'/api/services/pm/CommonParameter/GetAllShowNodes?'+'UUID='+robot_nodes_list[0]+'&Key=产品型号'
    # 启动程序时获取车型拿到LoadPositionX的值
    car_for_load_position_x = float(requests.get(robotune_get_all_show_nodes).json()["result"][0]["childNodes"][0]["parameterValue"])
    # 读取车体车型
    CAR_MODEL = requests.get(robotune_get_all_show_nodes_car_model).json()["result"][0]["childNodes"][0]["parameterValue"]
    print(f"成功读取到LoadPositionX：{car_for_load_position_x}")
    print(f"当前车型为：{CAR_MODEL}")
    # 根据模板复制一份excel文件，并返回文件路径
    file_path = get_copy_new_file("/home/visionnav/sim_res_bak/test_case/","/home/visionnav/sim_res_bak/test_result/get_put/","get_put_sim_case_temp")

    # 初始化测试用例
    read_excel_lines(file_path)

    # 订阅ecal发送的svc/pose消息
    ecal_core.initialize(sys.argv, "Python Protobuf Subscriber")
    sub = ProtoSubscriber("svc/pose", pose_pb2.Pose)
    # 方法用于设置消息回调函数，当收到消息时，会调用这个函数。
    sub.set_callback(on_message)

    # TODO: 如果是ST、SL车体中心在货叉挡板后面则应该+loadpositionx的值，最终得到挡板的姿态
    # TODO: 如果是P车、E车或者其他P车体中心在货叉挡板后面则应该-loadpositionx的值，最终得到挡板的姿态
    # TODO: 当前仿真世界坐标系的值是以卡板中心为准，所以需要优化下卡板中心到前端的偏移量
    # TODO: 已经知道当前的卡板长度、宽度、当前卡板的中心坐标，要求最终到车体中心距离是否一致
    # TODO: 获取车体运行的状态，尝试在取货阶段获取到感知检测的结果，分析感知的最大值，最小值和检测波动{'result': {'currentPower': '100%', 'runstatus': {'state': 2, 'maxVelocity': 3.0, 'maxVelocityLimitReason': 'ControlCenter', 'controlError': '100000.000,0.000,0.0000', 'lastEndControlError': '0.011,0.001,-0.0015', 'locationType': 3, 'locationMode': 0, 'chargeMode': 2, 'positionInitialized': 'true', 'taskData': {'source': 0, 'taskType': 0, 'taskId': 0, 'taskStatus': 0, 'timeStamp': '2025-04-03 02:09:51:2470'}, 'goodsData': {'hasGoods': 'false', 'poseDetected': '-0.074,0.003,0.0020,0.000,0.000,0.000', 'poseOnVehicle': '-2.082,-0.017,0.0060'}, 'lastUpdate': '2025-04-03 10:09:51:248'}}, 'targetUrl': None, 'success': True, 'error': None, 'unAuthorizedRequest': False, '__abp': True}
    while True:
        try:
            # response_GetFlowInfo = requests.get(robotune_task_GetFlowInfo).json()
            response_GetCurrentTaskInfo = requests.get(robotune_tail_get_task_info).json()
            if response_GetCurrentTaskInfo["result"] is not None:
                task_id = response_GetCurrentTaskInfo["result"]["taskId"]
                response_agv_debug = requests.get(f"{robotune_get_debug_status}{task_id}")
                robotune_mode = response_agv_debug.json()["result"]["flowName"]
                robotune_group_title = response_agv_debug.json()["result"]["taskGroup"]["groupName"]
                case_title = robotune_mode+robotune_group_title
                print("当前任务为："+response_GetCurrentTaskInfo["result"]["taskTypeName"])
                time.sleep(1)
                if response_GetCurrentTaskInfo["result"]["taskTypeName"] != "Stop":
                    case_title = None
                else:
                    IS_OPEN_ECAL_CALLBAK = True
            else:
                print(f"等待robotune任务下发")
                time.sleep(1)
        except requests.exceptions.ConnectionError:
            print(f"robotune程序未启用，请启动robotune程序,再执行此脚本")
            sys.exit(1)
        except KeyboardInterrupt:
            print("用户执行退出程序")
            ecal_core.finalize()
            sys.exit(0)
        except Exception as e:
            print(f"未知错误:{e}")
            sys.exit(1)
            